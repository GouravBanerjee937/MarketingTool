"""Incremental Analysis workbook — ONE header row + ONE data row.

Every attribute across all stages is its own column. A brand's whole run is a
single row that fills up as the user progresses (refresh() is called after each
stage action and rewrites that one row). Nothing is written as extra rows.
"""
import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.brand import Brand
from app.models.competitor import Competitor, CompetitorScope
from app.models.persona import Persona
from app.models.run import LlmRun

EXPORTS_DIR = Path(__file__).resolve().parents[2] / "exports"

HEADER_FILL = PatternFill("solid", fgColor="4F2EAA")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(vertical="top", wrap_text=True)

_STAGE_ORDER = {"competitors:fetch": 1, "competitors:analyze": 2, "content:themes": 3, "content:generate": 4}


def path_for(brand_id: uuid.UUID) -> Path:
    return EXPORTS_DIR / f"analysis-{brand_id}.xlsx"


def _yn(social: dict, key: str) -> str:
    return "yes" if isinstance(social, dict) and social.get(key) else "no"


def _stage_key(stage: str) -> int:
    for prefix, order in _STAGE_ORDER.items():
        if stage.startswith(prefix):
            return order
    return 9


async def build_bytes(db: AsyncSession, brand_id: uuid.UUID) -> bytes | None:
    brand = await db.get(Brand, brand_id)
    if brand is None:
        return None

    personas = (
        (
            await db.execute(
                select(Persona)
                .where(Persona.brand_id == brand_id)
                .options(selectinload(Persona.variants))
                .order_by(Persona.position, Persona.created_at)
            )
        )
        .scalars()
        .all()
    )
    competitors = (
        (
            await db.execute(
                select(Competitor)
                .where(Competitor.brand_id == brand_id)
                .order_by(Competitor.is_primary.desc(), Competitor.position, Competitor.name)
            )
        )
        .scalars()
        .all()
    )
    scope = (
        await db.execute(select(CompetitorScope).where(CompetitorScope.brand_id == brand_id))
    ).scalar_one_or_none()
    regions = ", ".join(scope.regions) if scope and scope.regions else "—"
    runs = (await db.execute(select(LlmRun).where(LlmRun.brand_id == brand_id))).scalars().all()
    runs = sorted(
        runs,
        key=lambda r: (_stage_key(r.stage), r.created_at or datetime.min.replace(tzinfo=timezone.utc), r.seq),
    )

    considered = [c for c in competitors if c.status == "considered"]
    targets = considered[:3]
    target_names = {c.name for c in targets}

    def persona_str(p: Persona) -> str:
        meta = ", ".join(x for x in (p.user_type, p.business_size, p.region) if x)
        variants = "; ".join(
            f"{v.label} — {v.description}" if v.description else v.label for v in p.variants
        )
        return (
            f"{p.name}" + (f" ({meta})" if meta else "")
            + f"\n  pain: {p.pain_points or '—'}"
            + f"\n  platforms: {p.current_platforms or '—'}"
            + f"\n  goal: {p.main_goal or '—'}"
            + (f"\n  variants: {variants}" if variants else "")
        )

    def comp_line(c: Competitor) -> str:
        return f"{c.name}{' [primary]' if c.is_primary else ''} — {c.source} — {c.website or '—'} — {c.description or '—'}"

    def analysis_str(c: Competitor) -> str:
        a = c.analysis or {}
        social = a.get("social") or {}
        feats = " | ".join(
            f"{f.get('feature', '')}: {f.get('sample_marketing', '')} (src: {f.get('source', 'NA')})"
            for f in (a.get("features") or [])
        )
        return (
            f"{c.name}:"
            f"\n  Revenue USD: {a.get('revenue_usd') or 'NA'}"
            f"\n  Revenue INR: {a.get('revenue_inr') or 'NA'} (src: {a.get('revenue_source') or 'NA'})"
            f"\n  Users: {a.get('users') or 'NA'} (src: {a.get('users_source') or 'NA'})"
            f"\n  Moats: {'; '.join(a.get('moats') or []) or 'NA'}"
            f"\n  Social: IG {_yn(social, 'instagram')}, Blog {_yn(social, 'blog')}, "
            f"FB {_yn(social, 'facebook')}, X {_yn(social, 'x')}, Other {social.get('thirdparty') or 'NA'}"
            f"\n  Features: {feats or 'NA'}"
        )

    def grp(prefix: str, names: set | None = None) -> tuple[str, str, str]:
        sysL, usrL, rspL = [], [], []
        for r in runs:
            if not r.stage.startswith(prefix):
                continue
            if prefix == "competitors:analyze:" and names is not None and r.stage[len(prefix):] not in names:
                continue
            sysL.append(f"[{r.stage}]\n{r.system_prompt}")
            usrL.append(f"[{r.stage}]\n{r.user_prompt}")
            rspL.append(f"[{r.stage}]\n{r.response}")
        j = lambda L: ("\n\n".join(L) or "—")  # noqa: E731
        return j(sysL), j(usrL), j(rspL)

    personas_cell = "\n\n".join(persona_str(p) for p in personas) or "—"
    considered_cell = "\n".join(comp_line(c) for c in competitors if c.status == "considered") or "—"
    review_cell = "\n".join(comp_line(c) for c in competitors if c.status == "pending") or "—"
    rejected_cell = "\n".join(comp_line(c) for c in competitors if c.status == "rejected") or "—"
    analysis_cell = "\n\n".join(analysis_str(c) for c in targets if c.analysis) or "—"
    fetch_sys, fetch_usr, fetch_rsp = grp("competitors:fetch")
    an_sys, an_usr, an_rsp = grp("competitors:analyze:", target_names)
    th_sys, th_usr, th_rsp = grp("content:themes")
    sc_sys, sc_usr, sc_rsp = grp("content:generate")

    columns: list[tuple[str, str]] = [
        ("Serial no.", "1"),
        ("Brand", brand.name),
        ("Vision", brand.vision or "—"),
        ("Goal", brand.goal or "—"),
        ("Moat", brand.moat or "—"),
        ("Operating regions", regions),
        ("Personas (ICP)", personas_cell),
        ("Competitors — considered (shortlist)", considered_cell),
        ("Competitors — to review", review_cell),
        ("Competitors — rejected", rejected_cell),
        ("Competitor analysis (max 3)", analysis_cell),
        ("Competitor fetch — system prompt", fetch_sys),
        ("Competitor fetch — user prompt (inputs)", fetch_usr),
        ("Competitor fetch — response (output)", fetch_rsp),
        ("Competitor analysis — system prompt", an_sys),
        ("Competitor analysis — user prompt (inputs)", an_usr),
        ("Competitor analysis — response (output)", an_rsp),
        ("Content themes — system prompt", th_sys),
        ("Content themes — user prompt (inputs)", th_usr),
        ("Content themes — response (output)", th_rsp),
        ("Content script — system prompt", sc_sys),
        ("Content script — user prompt (inputs)", sc_usr),
        ("Content script — response (output)", sc_rsp),
    ]

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Analysis"
    for ci, (name, val) in enumerate(columns, start=1):
        h = ws.cell(row=1, column=ci, value=name)
        h.font = HEADER_FONT
        h.fill = HEADER_FILL
        h.alignment = WRAP
        ws.column_dimensions[get_column_letter(ci)].width = 42
        ws.cell(row=2, column=ci, value=val).alignment = WRAP
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def refresh(db: AsyncSession, brand_id: uuid.UUID | None) -> None:
    if brand_id is None:
        return
    try:
        data = await build_bytes(db, brand_id)
        if data is None:
            return
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        path_for(brand_id).write_bytes(data)
    except Exception:  # noqa: BLE001 — reporting must never break a stage action
        pass
